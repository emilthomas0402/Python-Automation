import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time

# Set the Chrome driver path
chrome_options = Options()
chrome_options.add_argument("chromedriver-win64")  # Provide the path to your Chrome driver executable

# Initialize a web driver with ChromeOptions
driver = webdriver.Chrome(options=chrome_options)

# URL of the website you want to search
website_url = 'https://showtime_internal:hJ8cdQAm6fmhXJL@www.skyshowtime.com/staff-trial'  # Replace with the actual website URL

# Open the website
driver.get(website_url)

# Manually login to the application (You will need to fill in your login details manually)

# Wait for 5 seconds after login
time.sleep(5)  # Wait for 5 seconds (adjust as needed)

# Find the adult profile using the provided xpaths
try:
    # Find the adult profile using the provided XPath
    profile = WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.XPATH, "(//div[@class='profiles__avatar carousel-avatar profiles-main-page'])[1]")))
    profile.click()
    # print("Clicked on the adult profile.")
except TimeoutException:
    print("Failed to find the adult profile within 90 seconds.")
    driver.quit()

# Wait for 4 seconds for the home page to load
time.sleep(4)

# Code to work with Excel file
input_file = "inputSD.xlsx"
output_file = "Output.xlsx"

# Open the input sheet
input_workbook = openpyxl.load_workbook(input_file)
input_sheet = input_workbook.active

# Create output workbook and sheet
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Copy headers from the input sheet to the output sheet and add "Status" header
headers = []
for header in input_sheet[1]:
    headers.append(header.value)
headers.extend(["Status", ])  # Add "Status" header
output_sheet.append(headers)

# Iterate through each row in the input sheet
for idx, row in enumerate(input_sheet.iter_rows(min_row=2, values_only=True), start=2):
    deeplink = row[0]  # Assuming Deeplink is in the first column

    try:
        if deeplink:
            # Execute the deep link using Selenium
            driver.get(deeplink)

            # Wait for 10 seconds (you can adjust the waiting time if needed)
            time.sleep(8)

            # Check if the playback button exists (assuming this is a CSS selector)
            try:
                playback_button = driver.find_element(By.CSS_SELECTOR, 'svg.playback-button-icon.play > path')
                playback_button.click()

                # Wait for playback to start (adjust the waiting time based on the video loading time)
                time.sleep(25)  # Adjust this time based on the video loading time

                # Write status to the output sheet
                status = "Success"
                output_sheet.append(row + (status,))
                print(f"Asset {row[3]} (Row {idx}) played successfully.")

            except NoSuchElementException:
                status = "Failure"
                output_sheet.append(row + (status,))
                print(f"Failed to play asset {row[3]} (Row {idx}). Asset is not available in platform.")

        else:
            print(f"Input is None for Row {idx}. Skipping...")
    except Exception as e:
        status = "Failure"
        output_sheet.append(row + (status,))
        print(f"Failed to process asset {row[3]} (Row {idx}). Error: {str(e)}")

# Save the output workbook
output_workbook.save(output_file)
print("Output file generated successfully.")

# Close the Chrome driver
driver.quit()

#//*[@id="mount"]/div[2]/div/div/div[2]/div/div/div/div[2]/div[4]/div[2]/button[2]/svg